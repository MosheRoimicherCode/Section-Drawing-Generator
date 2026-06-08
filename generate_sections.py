from __future__ import annotations

import argparse
import math
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import ezdxf
from openpyxl import load_workbook

from config import (
    ARM_LAYERS,
    OUTPUT_DIR,
    POLES_FILE,
    SECTION_TEMPLATES,
    SURVEY_FILE,
    SURVEY_TO_OUTPUT_UNITS,
    TEMPLATES_DIR,
)


@dataclass(frozen=True)
class Pole:
    pole_number: str
    x: float
    y: float
    section_type: str


@dataclass(frozen=True)
class ArmPolyline:
    layer: str
    midpoint_x: float
    midpoint_y: float
    length_mm: int
    handle: str


@dataclass(frozen=True)
class AssignedArm:
    pole: Pole
    arm: ArmPolyline
    side: str


@dataclass(frozen=True)
class ReplacementResult:
    override_replacements: int
    rendered_replacements: int
    block_attribute_replacements: int
    missing_placeholders: list[str]
    non_numeric_dimension_values: list[str]


def read_poles(path: Path) -> list[Pole]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = {str(value).strip().lower(): index for index, value in enumerate(rows[0]) if value is not None}
    required = ["pole_number", "x", "y", "section_type"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Missing required poles.xlsx columns: {', '.join(missing)}")

    poles: list[Pole] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not row or row[headers["pole_number"]] in (None, ""):
            continue
        try:
            poles.append(
                Pole(
                    pole_number=str(row[headers["pole_number"]]).strip(),
                    x=float(row[headers["x"]]),
                    y=float(row[headers["y"]]),
                    section_type=str(row[headers["section_type"]]).strip(),
                )
            )
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid pole data on Excel row {row_number}: {row}") from exc
    return poles


def iter_polyline_vertices(entity) -> list[tuple[float, float, float]]:
    dxftype = entity.dxftype()
    if dxftype == "POLYLINE":
        return [(vertex.dxf.location.x, vertex.dxf.location.y, vertex.dxf.location.z) for vertex in entity.vertices]
    if dxftype == "LWPOLYLINE":
        elevation = float(entity.dxf.elevation or 0)
        return [(point[0], point[1], elevation) for point in entity.get_points("xy")]
    return []


def measure_3d_length(vertices: list[tuple[float, float, float]]) -> float:
    total = 0.0
    for start, end in zip(vertices, vertices[1:]):
        dx = end[0] - start[0]
        dy = end[1] - start[1]
        dz = end[2] - start[2]
        total += math.sqrt(dx * dx + dy * dy + dz * dz)
    return total


def midpoint(vertices: list[tuple[float, float, float]]) -> tuple[float, float]:
    if not vertices:
        raise ValueError("Cannot compute midpoint for an empty polyline")
    return (
        sum(point[0] for point in vertices) / len(vertices),
        sum(point[1] for point in vertices) / len(vertices),
    )


def read_arm_polylines(path: Path) -> list[ArmPolyline]:
    doc = ezdxf.readfile(path)
    msp = doc.modelspace()
    arms: list[ArmPolyline] = []

    for entity in msp.query("POLYLINE LWPOLYLINE"):
        layer = entity.dxf.layer
        if layer not in ARM_LAYERS:
            continue
        vertices = iter_polyline_vertices(entity)
        if len(vertices) < 2:
            continue
        mid_x, mid_y = midpoint(vertices)
        length_mm = round(measure_3d_length(vertices) * SURVEY_TO_OUTPUT_UNITS)
        arms.append(
            ArmPolyline(
                layer=layer,
                midpoint_x=mid_x,
                midpoint_y=mid_y,
                length_mm=length_mm,
                handle=entity.dxf.handle,
            )
        )
    return arms


def nearest_pole(arm: ArmPolyline, poles: Iterable[Pole]) -> Pole:
    return min(
        poles,
        key=lambda pole: (arm.midpoint_x - pole.x) ** 2 + (arm.midpoint_y - pole.y) ** 2,
    )


def assign_arms_to_poles(arms: list[ArmPolyline], poles: list[Pole]) -> list[AssignedArm]:
    assignments: list[AssignedArm] = []
    for arm in arms:
        pole = nearest_pole(arm, poles)
        side = "L" if arm.midpoint_x < pole.x else "R"
        assignments.append(AssignedArm(pole=pole, arm=arm, side=side))
    return assignments


def placeholder_ids(layer_config: object) -> tuple[str, ...]:
    if isinstance(layer_config, str):
        return (layer_config,)
    return tuple(str(value) for value in layer_config)


def assignment_distance_xy(assignment: AssignedArm) -> float:
    return math.sqrt(
        (assignment.arm.midpoint_x - assignment.pole.x) ** 2
        + (assignment.arm.midpoint_y - assignment.pole.y) ** 2
    )


def build_values_for_pole(pole: Pole, assignments: list[AssignedArm]) -> dict[str, str | int]:
    values: dict[str, str | int] = {
        "POLE_NUM": pole.pole_number,
        "POLE_NAME": pole.pole_number,
    }

    grouped_by_layer: dict[str, list[AssignedArm]] = defaultdict(list)
    grouped_by_layer_and_side: dict[tuple[str, str], list[AssignedArm]] = defaultdict(list)
    for assignment in assignments:
        if assignment.pole.pole_number == pole.pole_number:
            grouped_by_layer[assignment.arm.layer].append(assignment)
            grouped_by_layer_and_side[(assignment.arm.layer, assignment.side)].append(assignment)

    for layer, configured_placeholders in ARM_LAYERS.items():
        placeholders = placeholder_ids(configured_placeholders)
        if len(placeholders) == 1:
            candidates = sorted(grouped_by_layer.get(layer, []), key=assignment_distance_xy)
            if candidates:
                values[placeholders[0]] = candidates[0].arm.length_mm
            continue

        left_placeholder, right_placeholder = placeholders[:2]
        for side, placeholder in (("L", left_placeholder), ("R", right_placeholder)):
            candidates = sorted(grouped_by_layer_and_side.get((layer, side), []), key=assignment_distance_xy)
            if candidates:
                values[placeholder] = candidates[0].arm.length_mm

        for extra_index, placeholder in enumerate(placeholders[2:], start=2):
            candidates = sorted(grouped_by_layer.get(layer, []), key=assignment_distance_xy)
            if len(candidates) > extra_index:
                values[placeholder] = candidates[extra_index].arm.length_mm
    return values


def output_filename_for_pole(pole: Pole) -> str:
    return f"POLE_{pole.pole_number}_SECTION.dxf"


def missing_arm_problems(values: dict[str, str | int]) -> list[str]:
    problems: list[str] = []
    for layer, configured_placeholders in ARM_LAYERS.items():
        for placeholder in placeholder_ids(configured_placeholders):
            if placeholder not in values:
                problems.append(f"LINE MISSING: no closest polyline value found for layer {layer}, placeholder {placeholder}")
    return problems


def entity_text(entity) -> str:
    if entity.dxftype() == "TEXT":
        return entity.dxf.text or ""
    if entity.dxftype() == "MTEXT":
        return entity.plain_text() if hasattr(entity, "plain_text") else entity.text
    return ""


def set_entity_text(entity, value: str) -> None:
    if entity.dxftype() == "TEXT":
        entity.dxf.text = value
    elif entity.dxftype() == "MTEXT":
        if hasattr(entity, "set_content"):
            entity.set_content(value)
        else:
            entity.text = value


def replace_rendered_dimension_block_text(doc, dim, dimension_id: str, value: str) -> int:
    block_name = dim.dxf.get("geometry")
    if not block_name or block_name not in doc.blocks:
        return 0

    replacements = 0
    block = doc.blocks[block_name]
    for entity in block:
        if entity.dxftype() not in {"TEXT", "MTEXT"}:
            continue
        if entity_text(entity).strip() == dimension_id:
            set_entity_text(entity, value)
            replacements += 1
    return replacements


def is_as_made_text(value: str) -> bool:
    normalized = value.strip().upper().replace("-", " ")
    return normalized.startswith("AS MADE")


def replace_block_attribute_text(doc, block_name: str, attribute_tag: str, value: str) -> int:
    replacements = 0
    block_name_upper = block_name.upper()
    attribute_tag_upper = attribute_tag.upper()

    for layout in doc.layouts:
        for insert in layout.query("INSERT"):
            if insert.dxf.name.upper() != block_name_upper:
                continue

            for attrib in insert.attribs:
                if attrib.dxf.tag.upper() == attribute_tag_upper:
                    current_text = attrib.dxf.text or ""
                    if is_as_made_text(current_text):
                        attrib.dxf.text = value
                        replacements += 1

    return replacements


def replace_dimension_text(template_path: Path, output_path: Path, values: dict[str, str | int]) -> ReplacementResult:
    doc = ezdxf.readfile(template_path)
    msp = doc.modelspace()
    override_replacements = 0
    rendered_replacements = 0
    block_attribute_replacements = 0
    replaced_placeholders: set[str] = set()
    non_numeric_dimension_values: list[str] = []

    for dim in msp.query("DIMENSION"):
        placeholder = (dim.dxf.text or "").strip()
        if placeholder in values:
            value = str(values[placeholder])
            if placeholder != "POLE_NUM" and not value.isdecimal():
                non_numeric_dimension_values.append(f"{placeholder}={value}")
            dim.dxf.text = value
            override_replacements += 1
            rendered_replacements += replace_rendered_dimension_block_text(doc, dim, placeholder, value)
            replaced_placeholders.add(placeholder)

    if "POLE_NUM" in values:
        block_attribute_replacements = replace_block_attribute_text(
            doc,
            "ANOT0",
            "TEXT",
            f"AS MADE _ POLE NUMBER {values['POLE_NUM']}",
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc.saveas(output_path)
    missing_placeholders = [placeholder for placeholder in values if placeholder not in replaced_placeholders]
    return ReplacementResult(
        override_replacements=override_replacements,
        rendered_replacements=rendered_replacements,
        block_attribute_replacements=block_attribute_replacements,
        missing_placeholders=missing_placeholders,
        non_numeric_dimension_values=non_numeric_dimension_values,
    )


def write_run_report(
    report_path: Path,
    started_at: datetime,
    ended_at: datetime,
    section_reports: list[tuple[str, list[str]]],
) -> None:
    lines = [
        "Section Drawing Generator Run Report",
        f"Started: {started_at.isoformat(timespec='seconds')}",
        f"Ended:   {ended_at.isoformat(timespec='seconds')}",
        f"Timespan: {ended_at - started_at}",
        "",
    ]

    if not section_reports:
        lines.append("No sections were processed.")
    else:
        for section_name, problems in section_reports:
            lines.append(f"[{section_name}]")
            if problems:
                for problem in problems:
                    lines.append(f"- {problem}")
            else:
                lines.append("- OK: no problems found")
            lines.append("")

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(lines), encoding="utf-8")


def generate_sections(
    debug: bool = False,
    limit: int | None = None,
    survey_file: str | Path = SURVEY_FILE,
    poles_file: str | Path = POLES_FILE,
    templates_dir: str | Path = TEMPLATES_DIR,
    output_dir: str | Path = OUTPUT_DIR,
    section_templates: dict[str, str] | None = None,
) -> None:
    survey_file = Path(survey_file)
    poles_file = Path(poles_file)
    templates_dir = Path(templates_dir)
    output_dir = Path(output_dir)
    section_templates = section_templates or SECTION_TEMPLATES

    started_at = datetime.now()
    poles = read_poles(poles_file)
    arms = read_arm_polylines(survey_file)
    assignments = assign_arms_to_poles(arms, poles)
    section_poles = sorted(poles, key=lambda pole: pole.pole_number)
    section_reports: list[tuple[str, list[str]]] = []

    if debug:
        print(f"Loaded {len(poles)} poles")
        for pole in poles:
            print(f"POLE {pole.pole_number}: x={pole.x} y={pole.y} section_type={pole.section_type}")
        print(f"Loaded {len(arms)} arm polylines")
        for arm in arms:
            print(f"ARM {arm.handle}: layer={arm.layer} midpoint=({arm.midpoint_x:.3f}, {arm.midpoint_y:.3f}) length_mm={arm.length_mm}")
        for assignment in assignments:
            print(f"ASSIGN {assignment.arm.handle}: pole={assignment.pole.pole_number} side={assignment.side}")

    generated = 0
    for pole in section_poles[:limit]:
        section_name = pole.pole_number
        problems: list[str] = []
        template_name = section_templates.get(pole.section_type)
        if not template_name:
            problems.append(f"TEMPLATE MISSING: no template configured for section_type {pole.section_type}")
            section_reports.append((section_name, problems))
            print(f"Skipping {section_name}: no template for section_type {pole.section_type}")
            continue

        template_path = templates_dir / template_name
        if not template_path.exists():
            problems.append(f"TEMPLATE FILE MISSING: {template_path}")
            section_reports.append((section_name, problems))
            print(f"Skipping {section_name}: missing template {template_path}")
            continue

        values = build_values_for_pole(pole, assignments)
        problems.extend(missing_arm_problems(values))
        output_path = output_dir / output_filename_for_pole(pole)

        # Copy first so the output is a template-derived file even if future metadata is added before replacement.
        shutil.copyfile(template_path, output_path)
        replacement_result = replace_dimension_text(output_path, output_path, values)
        for placeholder in replacement_result.missing_placeholders:
            problems.append(f"DIMENSION ID MISSING IN TEMPLATE: {placeholder}")
        for value in replacement_result.non_numeric_dimension_values:
            problems.append(f"DIMENSION NOT A NUMBER: {value}")
        if replacement_result.block_attribute_replacements == 0:
            problems.append("BLOCK ATTRIBUTE MISSING: block ANOT0 attribute TEXT was not found for pole number")
        generated += 1
        section_reports.append((section_name, problems))

        if debug:
            print(f"VALUES {output_path.name}: {values}")
            print(
                f"WROTE {output_path} "
                f"({replacement_result.override_replacements} dimension overrides, "
                f"{replacement_result.rendered_replacements} rendered text replacements, "
                f"{replacement_result.block_attribute_replacements} ANOT0 TEXT attribute replacements)"
            )

    ended_at = datetime.now()
    report_path = output_dir / f"run_report_{started_at.strftime('%Y%m%d_%H%M%S')}.txt"
    write_run_report(report_path, started_at, ended_at, section_reports)

    print(f"Generated {generated} section DXF file(s) in {output_dir}")
    print(f"Wrote run report: {report_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate railway pole section DXF drawings from survey measurements.")
    parser.add_argument("--debug", action="store_true", help="Print loaded poles, measured arms, assignments, and output values.")
    parser.add_argument("--limit", type=int, default=None, help="Generate only the first N poles.")
    parser.add_argument("--survey", default=str(SURVEY_FILE), help="Survey DXF path.")
    parser.add_argument("--poles", default=str(POLES_FILE), help="Poles Excel path.")
    parser.add_argument("--templates-dir", default=str(TEMPLATES_DIR), help="Folder containing template DXF files.")
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR), help="Folder where output DXF and report files are written.")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    generate_sections(
        debug=args.debug,
        limit=args.limit,
        survey_file=args.survey,
        poles_file=args.poles,
        templates_dir=args.templates_dir,
        output_dir=args.output_dir,
    )
